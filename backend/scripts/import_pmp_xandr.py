"""
Importa um export do Xandr Curate (XLSX) para as tabelas `pmp_deals` e
`pmp_deals_delivery`.

Quando rodar
------------
Uma vez, na migração inicial: substituímos a planilha manual "HYPR Product
Performance" pela área admin de PMP no report center. Esse script popula
o histórico de delivery (`pmp_deals_delivery`) e cria os registros master
em `pmp_deals` com os campos derivados do nome do deal (customer, agency,
campanha, quarter). Os campos manuais (PI value, status, margem, owner)
ficam pra serem preenchidos via UI.

Idempotente — pode rodar de novo:
  • `pmp_deals_delivery` faz UPSERT por (deal_id, day). Reimportar o mesmo
    export sobrescreve os mesmos registros.
  • `pmp_deals` não sobrescreve campos manuais: usamos `ensure_deal_from_xandr`
    que só insere se o deal_id ainda não existe.

Como rodar
----------
Da raiz do repo, com creds da SA configuradas:

    python -m backend.scripts.import_pmp_xandr \\
        --xlsx /path/to/Export_2026-05-21_08-19.xlsx

Flags:
  --dry-run        não escreve no BQ, só mostra o que faria
  --created-by EMAIL   atribuído como created_by nos registros master
                       (default: 'migration@hypr.mobi')

Formato esperado
----------------
Sheet "report" com headers (case-sensitive):
  Curated Deal Insertion Order Name, Curated Deal Insertion Order Id,
  Curated Deal Id, Curated Deal Name, Curator Margin Type Name,
  Curator Margin Type Id, Buyer Seat Name, Buyer Seat ID, Buyer Seat Code,
  Day, Curator Tech Fees, Imps, Curator Total Cost, Curator Revenue,
  Viewable Imps, Curator Net Media Cost, Curator Margin, Clicks
"""

import argparse
import logging
import os
import sys
from datetime import date, datetime
from typing import List, Dict

# Suporte tanto pra `python -m backend.scripts.import_pmp_xandr`
# quanto `python backend/scripts/import_pmp_xandr.py`.
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import pmp_deals  # noqa: E402


logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("import_pmp_xandr")


EXPECTED_HEADERS = [
    "Curated Deal Insertion Order Name",
    "Curated Deal Insertion Order Id",
    "Curated Deal Id",
    "Curated Deal Name",
    "Curator Margin Type Name",
    "Curator Margin Type Id",
    "Buyer Seat Name",
    "Buyer Seat ID",
    "Buyer Seat Code",
    "Day",
    "Curator Tech Fees",
    "Imps",
    "Curator Total Cost",
    "Curator Revenue",
    "Viewable Imps",
    "Curator Net Media Cost",
    "Curator Margin",
    "Clicks",
]


def _parse_day(value) -> date:
    """Aceita date, datetime ou string ISO."""
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        return datetime.strptime(value[:10], "%Y-%m-%d").date()
    raise ValueError(f"Day inválido: {value!r}")


def read_xlsx(xlsx_path: str) -> List[Dict]:
    """Lê o sheet 'report' e devolve lista de dicts (uma linha por dia)."""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    if "report" not in wb.sheetnames:
        raise RuntimeError(
            f"Sheet 'report' não encontrado em {xlsx_path}. "
            f"Sheets disponíveis: {wb.sheetnames}"
        )
    ws = wb["report"]

    rows_iter = ws.iter_rows(values_only=True)
    headers = list(next(rows_iter))
    missing = [h for h in EXPECTED_HEADERS if h not in headers]
    if missing:
        raise RuntimeError(
            f"Headers obrigatórios ausentes no XLSX: {missing}. "
            f"Encontrados: {headers}"
        )

    idx = {h: headers.index(h) for h in EXPECTED_HEADERS}
    out = []
    for row in rows_iter:
        if row is None:
            continue
        if all(c is None for c in row):
            continue
        try:
            out.append({
                "io_name":           row[idx["Curated Deal Insertion Order Name"]],
                "io_id":             row[idx["Curated Deal Insertion Order Id"]],
                "deal_id":           str(row[idx["Curated Deal Id"]]),
                "curated_deal_name": row[idx["Curated Deal Name"]],
                "day":               _parse_day(row[idx["Day"]]),
                "tech_fees":         row[idx["Curator Tech Fees"]],
                "imps":              row[idx["Imps"]],
                "total_cost":        row[idx["Curator Total Cost"]],
                "revenue":           row[idx["Curator Revenue"]],
                "viewable_imps":     row[idx["Viewable Imps"]],
                "net_media_cost":    row[idx["Curator Net Media Cost"]],
                "margin":            row[idx["Curator Margin"]],
                "clicks":            row[idx["Clicks"]],
            })
        except Exception as e:
            log.warning(f"linha ignorada (parse falhou): {e} / row={row}")
    return out


def main():
    parser = argparse.ArgumentParser(description="Importa Xandr Curate XLSX → BigQuery (pmp_deals + pmp_deals_delivery).")
    parser.add_argument("--xlsx", required=True, help="caminho do Export_*.xlsx do Xandr")
    parser.add_argument("--dry-run", action="store_true", help="não escreve no BQ")
    parser.add_argument("--created-by", default="migration@hypr.mobi",
                         help="email atribuído como created_by/updated_by nos registros novos")
    args = parser.parse_args()

    if not os.path.exists(args.xlsx):
        log.error(f"Arquivo não encontrado: {args.xlsx}")
        sys.exit(1)

    log.info(f"Lendo {args.xlsx}")
    rows = read_xlsx(args.xlsx)
    log.info(f"{len(rows)} linhas lidas do XLSX")

    # Agrupa por deal pra mostrar resumo + criar master
    deals_seen: Dict[str, Dict] = {}
    for r in rows:
        d = deals_seen.setdefault(r["deal_id"], {
            "deal_id":           r["deal_id"],
            "curated_deal_name": r["curated_deal_name"],
            "io_name":           r["io_name"],
            "days":              0,
            "imps":              0.0,
            "revenue":           0.0,
            "cost":              0.0,
        })
        d["days"]    += 1
        d["imps"]    += float(r["imps"] or 0)
        d["revenue"] += float(r["revenue"] or 0)
        d["cost"]    += float(r["total_cost"] or 0)

    log.info(f"{len(deals_seen)} deals únicos:")
    for d in sorted(deals_seen.values(), key=lambda x: x["curated_deal_name"] or ""):
        parsed = pmp_deals.parse_deal_name(d["curated_deal_name"] or "")
        margin_pct = ((d["revenue"] - d["cost"]) / d["revenue"] * 100) if d["revenue"] else 0
        log.info(
            f"  • {d['deal_id']:>10} ({d['days']:>2}d)  "
            f"customer={parsed['customer'] or '?':<15} "
            f"campaign={(parsed['campaign_name'] or '?')[:30]:<30} "
            f"quarter={parsed['flight_quarter'] or '?':<10}  "
            f"rev={d['revenue']:>10,.2f}  margin={margin_pct:>5.1f}%"
        )

    if args.dry_run:
        log.info("--dry-run: nada gravado.")
        return

    # 1) Garante schema
    log.info("Criando tabelas (idempotente)...")
    schema_res = pmp_deals.setup_schema()
    log.info(f"  setup_schema → {schema_res}")

    # 2) Cria registros master (skip se já existem)
    created_count = 0
    for d in deals_seen.values():
        created = pmp_deals.ensure_deal_from_xandr(
            deal_id=d["deal_id"],
            curated_deal_name=d["curated_deal_name"] or "",
            io_name=d["io_name"],
            created_by=args.created_by,
        )
        if created:
            created_count += 1
    log.info(f"Master: {created_count} criados, {len(deals_seen) - created_count} já existiam")

    # 3) Upsert delivery
    delivery_rows = [{
        "deal_id":                str(r["deal_id"]),
        "day":                    r["day"],
        "imps":                   int(r["imps"] or 0),
        "viewable_imps":          int(r["viewable_imps"] or 0),
        "clicks":                 int(r["clicks"] or 0),
        "curator_net_media_cost": float(r["net_media_cost"] or 0),
        "curator_tech_fees":      float(r["tech_fees"] or 0),
        "curator_total_cost":     float(r["total_cost"] or 0),
        "curator_revenue":        float(r["revenue"] or 0),
        "curator_margin":         float(r["margin"] or 0),
    } for r in rows]

    log.info(f"Upsert de {len(delivery_rows)} linhas de delivery...")
    res = pmp_deals.upsert_delivery_rows(delivery_rows)
    log.info(f"  → {res}")

    log.info("Pronto.")


if __name__ == "__main__":
    main()
